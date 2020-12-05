import openpyxl
import random
from openpyxl import Workbook, load_workbook
import glob
import json


def D3_wellness_dialog_for_autoregressive():
  root_path = "./TK_data/T1_wellness"
  root_out = "./TK_data/T0_data"
  wellness_answer_file = root_path + "/T1D1_wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/T1D2_wellness_dialog_question.txt"
  wellness_autoregressive_file = root_out + "/T1_wellness_dialog_for_autoregressive.txt"


  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write(ques_data[1][:-1]+'\n')
        autoregressive_file.write(ans_data[1])
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()




def D1_tweet_dialog_dataset():
  root_path = "./TK_data/T5_twitter"
  root_out = "./TK_data/T0_data"
  tweet_file = root_path + "/T5_twitter.xlsx"
  tweet_file_output = root_out + "/T5_tweeter_"


  wb = load_workbook(filename=tweet_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  number = 1
  for row in ws.iter_rows():
    tweet_file_output_now = tweet_file_output + str(number).zfill(4) + '.txt'
    print("see this {}".format(tweet_file_output_now))
    f = open(tweet_file_output_now, 'w')
    for cell in row:
      if cell.value == None:
        break
      f.write(cell.value + "\n")
    f.close()
    number += 1

def D1_national_dataset():
    FROM ="./TK_data/T3_national"
    TO ="./TK_data/T0_data/T3_national_"

    number=1
    for json_path in glob.glob(FROM+"/JSON/*.json"):
        txt_path = TO + str(number).zfill(4) +'.txt'
        print("see this {}".format(txt_path))

        with open(json_path, 'r') as json_file :
            py_data = json.load(json_file)
            py_document = py_data['document'][0]
            py_utterance = py_document['utterance']
            with open(txt_path, 'w') as txt_file :
                for utterance in py_utterance :
                    txt_file.write(utterance['form']+'\n')
        number += 1
if __name__ == "__main__":
  #D3_wellness_dialog_for_autoregressive()

  D1_tweet_dialog_dataset()
  D1_national_dataset()
