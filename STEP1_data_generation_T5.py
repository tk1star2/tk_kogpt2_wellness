import openpyxl
import random
from openpyxl import Workbook, load_workbook

def D1_tweet_dialog_dataset():
  root_path = "./TK_data/T5_twitter"
  tweet_file = root_path + "/T5_twitter.xlsx"
  tweet_file_output = root_path + "/T5_tweeter.txt"

  f = open(tweet_file_output, 'w')

  wb = load_workbook(filename=tweet_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  for row in ws.iter_rows():
    for cell in row:
      if cell.value == None:
        break
      # print(cell.value)
      f.write(cell.value + "\n")
    # print("\n\n\n")
    f.write("\n\n\n")

  f.close()

def D2_tweet_data_for_autoregressive():
  root_path = "./TK_data/T5_twitter"

  file_path = root_path + "/T5_tweeter.txt"
  tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive.txt"

  data_file = open(file_path, 'r')
  tweet_file = open(tweeter_autoregressive_file, 'w')

  data_file_lines = data_file.readlines()
  dialog = ''
  for line_num, line_data in enumerate(data_file_lines):
    if line_data == "\n" and dialog != '':
      dialog += "\n"
      tweet_file.write(dialog)
      #print(dialog)
      dialog = ''
    elif line_data != "\n":
      dialog += "<s>" + line_data[:-1] + "</s>"
  data_file.close()
  tweet_file.close()

def tweeter_autoregressive_data_with_token():
  root_path = "./TK_data/T5_twitter"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive_with_token.txt"

  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write("<s>" + ques_data[1][:-1] + "</s><s>" + ans_data[1][:-1] + "</s>\n")
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()

def category_data():
  root_path = "./TK_data/T5_twitter"
  data_path = root_path + "/chatbot_wellness_data.txt"
  c_output = root_path + "/chatbot_wellness_category.txt"

  i_f = open(data_path, 'r')
  o_f = open(c_output, 'w')

  category_count = 0
  flag = True

  cate_dict = []
  i_lines = i_f.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    a = tmp[1][:-1]
    q = tmp[0]
    if a not in cate_dict:
      cate_dict.append(a)
      o_f.write(a.strip() + "    " + str(category_count) + "\n")
      category_count += 1
  o_f.close()
  i_f.close()

def merge_data():
  root_path = "./TK_data/T5_twitter"

  chatbot_file = root_path + "/chatbot_data.txt"
  wellness_file = root_path + "/wellness_dialog.txt"

  total_data_file = root_path + "/a_chatbot_wellness_data.txt"

  chatbot_f = open(chatbot_file, 'r')
  wellness_f = open(wellness_file, 'r')
  output_f = open(total_data_file, 'w')

  chatbot_lines = chatbot_f.readlines()
  for line_num, line_data in enumerate(chatbot_lines):
    output_f.write(line_data)

  wellness_lines = wellness_f.readlines()
  for line_num, line_data in enumerate(wellness_lines):
    output_f.write(line_data)

  chatbot_f.close()
  wellness_f.close()
  output_f.close()

if __name__ == "__main__":
  '''
  root_path = "./TK_data/T5_twitter"
  file_path = root_path + "/chatbot_wellness_data.txt"
  o_path = root_path + "/chatbot_wellness_data_for_autoregressive.txt"

  i_file = open(file_path, 'r')
  o_file = open(o_path, 'w')

  i_lines = i_file.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    question = tmp[0]
    answer = tmp[1][:-1]
    o_file.write("<s>" + question + "</s><s>" + answer+ "</s>\n")
  '''
  
  D1_tweet_dialog_dataset()
  D2_tweet_data_for_autoregressive()

  #wellness_answer_data()
  #wellness_question_data()
  #wellness_dialog_for_autoregressive()
  #seperate_wellness_data()
