import openpyxl
import random
from openpyxl import Workbook, load_workbook

def D1_wellness_answer_data():
  root_path = "./TK_data/T1_wellness"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_a_output = root_path + "/T1D1_wellness_dialog_answer.txt"

  f = open(wellness_a_output, 'w')
  wb = load_workbook(filename=wellness_file)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows():
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[2].value + "\n")
  f.close()
def D2_wellness_question_data():
  root_path = "./TK_data/T1_wellness"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_q_output = root_path + "/T1D2_wellness_dialog_question.txt"

  f = open(wellness_q_output, 'w')

  wb = load_workbook(filename=wellness_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  for row in ws.iter_rows():
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()





def D3_wellness_dialog_for_autoregressive():
  root_path = "./TK_data/T1_wellness"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "/T1D1_wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/T1D2_wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "/T1D3_wellness_dialog_for_autoregressive.txt"


  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1])
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()


def D4_seperate_wellness_data():
  root_path = "./TK_data/T1_wellness"
  # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
  # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
  file_path = root_path + "/T1D3_wellness_dialog_for_autoregressive.txt"
  train_file_path = root_path + "/T1_wellness_train.txt"
  test_file_path = root_path + "/T1_wellness_test.txt"

  sperated_file = open(file_path, 'r')
  train_file = open(train_file_path, 'w')
  test_file = open(test_file_path, 'w')

  sperated_file_lines = sperated_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(sperated_file_lines):
    rand_num = random.randint(0, 10)
    if rand_num < 10:
      train_file.write(line_data)
    else:
      test_file.write(line_data)

  sperated_file.close()
  train_file.close()
  test_file.close()
def category_data():
  root_path = "./TK_data/T1_wellness"
  data_path = root_path + "/chatbot_wellness_data.txt"
  c_output = root_path + "/chatbot_wellness_category.txt"

  i_f = open(data_path, 'r')
  o_f = open(c_output, 'w')

  category_count = 0
  flag = True

  cate_dict = []
  i_lines = i_f.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    a = tmp[1][:-1]
    q = tmp[0]
    if a not in cate_dict:
      cate_dict.append(a)
      o_f.write(a.strip() + "    " + str(category_count) + "\n")
      category_count += 1
  o_f.close()
  i_f.close()

def wellness_text_classification_data():
  root_path = "./TK_data/T1_wellness"
  wellness_category_file = root_path + "/wellness_dialog_category.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"

  cate_file = open(wellness_category_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  text_classfi_file = open(wellness_text_classification_file, 'w')

  category_lines = cate_file.readlines()
  cate_dict = {}
  for line_num, line_data in enumerate(category_lines):
    data = line_data.split('    ')
    cate_dict[data[0]] = data[1][:-1]
  print(cate_dict)

  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    data = line_data.split('    ')
    # print(data[1]+ "    " + cate_dict[data[0]])
    text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

  cate_file.close()
  ques_file.close()
  text_classfi_file.close()

def merge_data():
  root_path = "./TK_data/T1_wellness"

  chatbot_file = root_path + "/chatbot_data.txt"
  wellness_file = root_path + "/wellness_dialog.txt"

  total_data_file = root_path + "/a_chatbot_wellness_data.txt"

  chatbot_f = open(chatbot_file, 'r')
  wellness_f = open(wellness_file, 'r')
  output_f = open(total_data_file, 'w')

  chatbot_lines = chatbot_f.readlines()
  for line_num, line_data in enumerate(chatbot_lines):
    output_f.write(line_data)

  wellness_lines = wellness_f.readlines()
  for line_num, line_data in enumerate(wellness_lines):
    output_f.write(line_data)

  chatbot_f.close()
  wellness_f.close()
  output_f.close()

if __name__ == "__main__":
  '''
  root_path = "./TK_data/T1_wellness"
  file_path = root_path + "/chatbot_wellness_data.txt"
  o_path = root_path + "/chatbot_wellness_data_for_autoregressive.txt"

  i_file = open(file_path, 'r')
  o_file = open(o_path, 'w')

  i_lines = i_file.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    question = tmp[0]
    answer = tmp[1][:-1]
    o_file.write("<s>" + question + "</s><s>" + answer+ "</s>\n")
  '''
  
  D1_wellness_answer_data()
  D2_wellness_question_data()
  D3_wellness_dialog_for_autoregressive()
  D4_seperate_wellness_data()
